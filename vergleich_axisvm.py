# -*- coding: utf-8 -*-
"""
vergleich_axisvm.py
---------------------------------------------------------------------------
ABGLEICH DES WERKZEUGS GEGEN EIN AXISVM-STABMODELL, Stelle für Stelle.

    python3 vergleich_axisvm.py <export.xlsx> <vergleich_werkzeug.json>

<export.xlsx>  die Ausgabe aus AxisVM: die Blätter «Knoten», «Stäbe» und je
               Lastfall ein Blatt «vm <Name>» mit den Randspannungen.
<...json>      die Werkzeugseite, geschrieben von vergleich_werkzeug.mjs.

WAS DAS WERKZEUG SELBST FINDET
Die Zuordnung der Stäbe (Obergurt links, Vertikalblech, …) wird aus der
GEOMETRIE erschlossen, nicht eingegeben: die Gurte laufen in Jochachse und
liegen auf zwei Höhen, die Bleche stehen quer dazu. Ebenso der Versatz
zwischen den beiden Koordinatensystemen - er folgt aus dem x-Bereich der
Gurte.

Die Zuordnung der LASTFÄLLE wird geraten (Name gegen Name) und ausgewiesen;
mit --lastfall "vm snow=sk" lässt sie sich setzen.

Nichts an diesem Werkzeug ist an ein bestimmtes Bauwerk gebunden.
---------------------------------------------------------------------------
"""
import sys, json, math, collections, argparse

try:
    import openpyxl
except ImportError:
    sys.exit('Es fehlt openpyxl:  pip3 install openpyxl')


# --- 1. Das AxisVM-Modell lesen --------------------------------------------
def lies_modell(wb):
    """Knoten und Stäbe aus dem Export."""
    kn, st = {}, {}
    for r in wb['Knoten'].iter_rows(min_row=2, values_only=True):
        if r[0] is None or not isinstance(r[0], (int, float)):
            continue
        kn[int(r[0])] = (float(r[1]), float(r[2]), float(r[3]))
    for r in wb['Stäbe'].iter_rows(min_row=2, values_only=True):
        if r[0] is None or not isinstance(r[0], (int, float)):
            continue
        st[int(r[0])] = dict(i=int(r[1]), j=int(r[2]), L=float(r[3]))
    return kn, st


def klassiere(kn, st, geo):
    """Jeden Stab einer Bauteilfamilie zuordnen - aus der Geometrie.

    Die Gurte laufen in Jochachse. Ihre beiden Höhenlagen sind die z-Werte,
    auf denen die meisten dieser Stäbe liegen; welcher der obere ist, sagt
    das Vorzeichen. Quer dazu stehen die Bleche: lotrecht die
    Vertikalbleche, waagrecht die Horizontalbleche.
    """
    laengs = []
    for n, s in st.items():
        a, b = kn[s['i']], kn[s['j']]
        d = (b[0] - a[0], b[1] - a[1], b[2] - a[2])
        if abs(d[0]) > 1e-6 and abs(d[2]) < 1e-4:
            laengs.append((n, (a[2] + b[2]) / 2, (a[1] + b[1]) / 2))
    if not laengs:
        sys.exit('Im Export läuft kein Stab in Jochachse - ist es ein Tragjoch?')

    # Die zwei Hoehenlagen: haeufigste z-Werte, auf 5 mm gerundet
    haeuf = collections.Counter(round(z, 3) for _, z, _ in laengs)
    zwei = sorted(z for z, _ in haeuf.most_common(2))
    if len(zwei) < 2:
        sys.exit('Es liessen sich keine zwei Gurthöhen finden.')
    zUG, zOG = zwei
    zAchse = (zOG + zUG) / 2
    hAx = zOG - zUG

    # Der Versatz: die Gurte beginnen im Werkzeug bei x = 0.
    xs = [min(kn[st[n]['i']][0], kn[st[n]['j']][0]) for n, _, _ in laengs]
    versatz = min(xs)

    kl = {}
    for n, s in st.items():
        a, b = kn[s['i']], kn[s['j']]
        dx, dy, dz = (b[0] - a[0], b[1] - a[1], b[2] - a[2])
        zm, ym = (a[2] + b[2]) / 2, (a[1] + b[1]) / 2
        if abs(dx) > 1e-6 and abs(dz) < 1e-4:
            fam = None
            if abs(zm - zOG) < 0.15: fam = 'OG'
            elif abs(zm - zUG) < 0.15: fam = 'UG'
            if fam:
                kl[n] = f'{fam}{"R" if ym > 0 else "L"}'
                continue
        if abs(dx) < 1e-6 and abs(dy) < 1e-6 and abs(dz) > 1e-6:
            # lotrecht: Blech nur, wenn es zwischen den beiden Gurthoehen liegt
            kl[n] = 'BLV' if zUG - 0.1 <= zm <= zOG + 0.1 else 'MAST'
            continue
        if abs(dx) < 1e-6 and abs(dz) < 1e-4 and abs(dy) > 1e-6:
            kl[n] = 'BLH'
            continue
        kl[n] = 'sonst'
    return kl, dict(zOG=zOG, zUG=zUG, zAchse=zAchse, h=hAx, versatz=versatz)


# --- 2. Die Spannungen je Lastfall -----------------------------------------
def lies_spannungen(wb, kn, st, kl, blatt):
    """Randspannungen je Stab und Stelle, auf die Jochachse bezogen."""
    ws = wb[blatt]
    punkte = collections.defaultdict(list)
    stab = None
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is not None and isinstance(r[0], (int, float)):
            stab = int(r[0])
        if r[3] is None or isinstance(r[3], str):
            continue
        s = st.get(stab)
        if s is None or stab not in kl:
            continue
        a = float(r[3])
        xi, xj = kn[s['i']][0], kn[s['j']][0]
        x = xi + (xj - xi) * (a / s['L'] if s['L'] else 0)
        z = lambda i: float(r[i] or 0)
        punkte[kl[stab]].append(dict(
            x=x, stab=stab, a=a,
            smin=z(5), smax=z(6),          # Randspannungen
            sv=max(abs(z(9)), abs(z(10))),  # Vergleichsspannung
        ))
    for f in punkte:
        punkte[f].sort(key=lambda p: p['x'])
    return punkte


def blechmoment(p, breite, dicke):
    """Rechteckblech: aus den Randspannungen zurück auf N und M."""
    W = dicke * breite ** 2 / 6.0                    # mm3
    return abs(p['smax'] - p['smin']) / 2 * W / 1e6  # kNm


# --- 3. Der Vergleich -------------------------------------------------------
def naechster(liste, x, tol):
    tr = [p for p in liste if abs(p['x'] - x) < tol]
    return tr


def hauptsache(argv=None):
    ap = argparse.ArgumentParser(add_help=True, description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('export'), ap.add_argument('werkzeug')
    ap.add_argument('--lastfall', action='append', default=[],
                    help='Zuordnung "vm snow=sk", mehrfach erlaubt')
    ap.add_argument('--stationen', action='store_true',
                    help='jede Station einzeln zeigen, nicht nur die Zusammenfassung')
    ap.add_argument('--rand', type=float, default=0.5,
                    help='so viel am Jochende auslassen [m], Vorgabe 0.5')
    a = ap.parse_args(argv)

    wz = json.load(open(a.werkzeug))
    if wz.get('format') != 'tragjoch-vergleich':
        sys.exit('Die zweite Datei stammt nicht von vergleich_werkzeug.mjs.')
    geo = wz['geometrie']
    wb = openpyxl.load_workbook(a.export, read_only=True, data_only=True)
    kn, st = lies_modell(wb)
    kl, ax = klassiere(kn, st, geo)

    print(f"Werkzeug : {wz['name']}  ·  {geo['typ']}  ·  L = {geo['L']} m")
    print(f"AxisVM   : {a.export}")
    print(f"Zuordnung: Gurthöhen z = {ax['zUG']:.3f} / {ax['zOG']:.3f} m  →  "
          f"h = {ax['h']*1000:.0f} mm (Werkzeug {geo['h']*1000:.0f} mm)")
    print(f"           x-Versatz {ax['versatz']:+.3f} m")
    dh = abs(ax['h'] - geo['h']) / geo['h'] if geo['h'] else 0
    if dh > 0.05:
        print(f"           ACHTUNG: die Gurtabstände weichen um {100*dh:.1f} % ab - "
              "vergleichen Sie zuerst die Geometrie.")
    zaehl = collections.Counter(kl.values())
    print('           Stäbe: ' + ' · '.join(f'{k} {v}' for k, v in sorted(zaehl.items())))
    print()

    # Lastfaelle zuordnen
    vm = [s for s in wb.sheetnames if s.lower().startswith('vm ')]
    fest = dict(z.split('=', 1) for z in a.lastfall)
    schluessel = {k: v['bez'] for k, v in wz['faelle'].items()}

    def raten(blatt):
        if blatt in fest: return fest[blatt]
        n = blatt[3:].lower()
        wort = {'self weight': 'gk', 'added weight': 'ak', 'snow': 'sk',
                'schnee': 'sk', 'eigengewicht': 'gk'}
        for k, v in wort.items():
            if k in n and v in wz['faelle']: return v
        if 'wind' in n:
            if '+y' in n or 'quer' in n:
                return 'wyk' if 'wyk' in wz['faelle'] else None
            if '+x' in n or 'längs' in n or 'laengs' in n:
                return 'wxk' if 'wxk' in wz['faelle'] else None
        return None

    tol = 0.25
    for blatt in vm:
        key = raten(blatt)
        if key is None:
            print(f'--- {blatt}: keine Entsprechung im Werkzeug '
                  f'(mit --lastfall "{blatt}=<key>" setzen; '
                  f'vorhanden: {", ".join(schluessel)})')
            continue
        sp = lies_spannungen(wb, kn, st, kl, blatt)
        stat = wz['faelle'][key]['stationen']
        print(f'=== {blatt}  ↔  {key} · {schluessel[key]} ===')
        if a.stationen:
            print(f"{'x':>7} | {'OG Wz':>8} {'Axis':>8} {'d':>7} | "
                  f"{'UG Wz':>8} {'Axis':>8} {'d':>7} | "
                  f"{'BLV Wz':>7} {'Axis':>7} {'d':>7}")
        sam = collections.defaultdict(list)
        for k in stat:
            xa = k['x'] + ax['versatz']
            if xa < ax['versatz'] + a.rand or xa > ax['versatz'] + geo['L'] - a.rand:
                continue
            werte = {}
            for fam, ids in (('OG', ('OG_L', 'OG_R')), ('UG', ('UG_L', 'UG_R'))):
                w = max((k['gurt'][i]['sig'] for i in ids if i in k['gurt']), default=0)
                p = [q['sv'] for f in (fam + 'L', fam + 'R')
                     for q in naechster(sp.get(f, []), xa, tol)]
                werte[fam] = (w, max(p) if p else 0)
            for fam, art in (('BLV', 'vertikal'), ('BLH', 'horizontal')):
                bl = [v for v in k['blech'].values() if v['art'] == art]
                w = max((v['M'] for v in bl), default=0)
                br = bl[0]['breite'] if bl else 0
                di = bl[0]['dicke'] if bl else 0
                p = [blechmoment(q, br, di) for q in naechster(sp.get(fam, []), xa, tol)]
                werte[fam] = (w, max(p) if p else 0)
            for fam, (w, p) in werte.items():
                gr = 2.0 if fam in ('OG', 'UG') else 0.05
                if p > gr:
                    sam[fam].append(w / p - 1)
            if a.stationen:
                def d(t, gr):
                    return (f'{(t[0]/t[1]-1)*100:+6.1f}%'
                            if t[1] > gr else '      -')
                print(f"{xa:>7.2f} | {werte['OG'][0]:>8.2f} {werte['OG'][1]:>8.2f} "
                      f"{d(werte['OG'], 2.0):>7} | {werte['UG'][0]:>8.2f} "
                      f"{werte['UG'][1]:>8.2f} {d(werte['UG'], 2.0):>7} | "
                      f"{werte['BLV'][0]:>7.3f} {werte['BLV'][1]:>7.3f} "
                      f"{d(werte['BLV'], 0.05):>7}")
        for fam in ('OG', 'UG', 'BLV', 'BLH'):
            v = sam[fam]
            if not v:
                print(f'   {fam:4s} keine vergleichbaren Werte')
                continue
            v.sort()
            med = v[len(v) // 2]
            print(f'   {fam:4s} Mittel {100*sum(v)/len(v):+6.1f} %   '
                  f'Median {100*med:+6.1f} %   '
                  f'Spanne {100*v[0]:+6.1f} … {100*v[-1]:+6.1f} %   '
                  f'({len(v)} Stationen)')
        print()


if __name__ == '__main__':
    hauptsache()
